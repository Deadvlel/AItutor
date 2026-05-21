import json
import os
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from dependencies import get_current_user
from models.course import chuDe, taiLieu, cauHoi, lichSuLamKT, cauTraLoi, aiLog
from services.ai_service import hoi_gia_su, tim_sgk, _xay_dung_ngu_canh, tao_de_thi_json
from services.upload_service import xu_ly_excel, xu_ly_pdf, xu_ly_word

router = APIRouter()
ID_DO_BAI = 2


# ── Schema ──────────────────────────────────────────────────────
class ChamDoBaiRequest(BaseModel):
    id_cau_hoi:  int | None = None   # None nếu dùng câu hỏi sinh từ Chroma
    cau_hoi:     str | None = None   # câu hỏi gốc (khi không có id)
    dap_an_mau:  str | None = None   # đáp án mẫu (khi không có id)
    cau_tra_loi: str


class SinhCauHoiRequest(BaseModel):
    id_tai_lieu: int         # id bài học trong bảng taiLieu
    so_cau:      int = 5     # số câu muốn sinh


class LuuKetQuaRequest(BaseModel):
    id_tai_lieu: int
    tong_cau:    int
    so_cau_dung: int


# ── Sinh câu hỏi từ Chroma (không cần upload) ───────────────────
@router.post("/sinh-cau-hoi")
def sinh_cau_hoi(
    req: SinhCauHoiRequest,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Sinh câu hỏi dò bài từ nội dung SGK trong ChromaDB."""
    # Lấy thông tin bài học
    bai = db.query(taiLieu).filter(taiLieu.id_taiLieu == req.id_tai_lieu).first()
    if not bai:
        raise HTTPException(404, "Không tìm thấy bài học")

    ten_mon = None
    if bai.id_chuDe:
        cd = db.query(chuDe).filter(chuDe.id_chuDe == bai.id_chuDe).first()
        if cd:
            ten_mon = cd.ten_chuDe

    # Lấy nội dung từ Chroma
    doan_list = tim_sgk(bai.tieuDe, ten_mon=ten_mon, top_k=5)
    ngu_canh  = _xay_dung_ngu_canh(doan_list)

    so_cau = max(3, min(10, req.so_cau))

    prompt = f"""Bạn là giáo viên Việt Nam. Dựa vào nội dung SGK sau, hãy tạo {so_cau} câu hỏi dò bài ngắn gọn.

NỘI DUNG SGK BÀI "{bai.tieuDe}":
{ngu_canh}

Trả về JSON thuần, không markdown, không giải thích:
{{
  "cau_hois": [
    {{
      "thu_tu": 1,
      "cau_hoi": "Câu hỏi ngắn gọn, rõ ràng bằng tiếng Việt?",
      "dap_an_mau": "Đáp án đầy đủ các ý chính."
    }}
  ]
}}

Yêu cầu:
- Câu hỏi ngắn, học sinh có thể trả lời bằng miệng trong 1-2 câu
- Bao phủ các khái niệm quan trọng của bài
- Không hỏi về số trang, không hỏi câu quá khó"""

    try:
        raw = tao_de_thi_json(prompt)
        raw = raw.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        data = json.loads(raw)
        cau_hois = data.get("cau_hois", [])
        if not cau_hois:
            raise ValueError("Không có câu hỏi")
    except Exception as e:
        # Fallback: sinh câu hỏi đơn giản từ tiêu đề
        cau_hois = [
            {
                "thu_tu": 1,
                "cau_hoi": f"Em hãy trình bày khái niệm chính của bài '{bai.tieuDe}'?",
                "dap_an_mau": "Học sinh trình bày được nội dung cơ bản của bài."
            }
        ]

    return {
        "tieu_de": bai.tieuDe,
        "ten_mon": ten_mon or "",
        "cau_hois": [
            {
                "id": None,
                "thu_tu": c.get("thu_tu", i + 1),
                "cau_hoi": c.get("cau_hoi", ""),
                "dap_an_mau": c.get("dap_an_mau", ""),
            }
            for i, c in enumerate(cau_hois)
        ]
    }


# ── Chấm điểm (hỗ trợ cả câu có id và câu sinh từ Chroma) ──────
@router.post("/cham-diem")
def cham_diem(
    req: ChamDoBaiRequest,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # Lấy câu hỏi + đáp án
    if req.id_cau_hoi:
        cau = db.query(cauHoi).filter(cauHoi.id_cauHoi == req.id_cau_hoi).first()
        if not cau:
            raise HTTPException(404, "Không tìm thấy câu hỏi")
        cau_hoi_text = cau.noiDung
        dap_an_text  = cau.dapAnMau
        goi_y_text   = cau.goiY
    elif req.cau_hoi and req.dap_an_mau:
        cau_hoi_text = req.cau_hoi
        dap_an_text  = req.dap_an_mau
        goi_y_text   = None
    else:
        raise HTTPException(400, "Thiếu thông tin câu hỏi")

    prompt = f"""Em là gia sư đang chấm bài dò bài.

Câu hỏi: {cau_hoi_text}
Đáp án chuẩn: {dap_an_text}
Câu trả lời học sinh: {req.cau_tra_loi}

Hãy đánh giá và trả về JSON thuần (không markdown, không giải thích thêm):
{{
  "ket_qua": "dung",
  "nhan_xet": "Nhận xét ngắn gọn bằng tiếng Việt, xưng Thầy gọi Em"
}}

Quy tắc ket_qua:
- "dung"     : trả lời đúng và đủ ý chính
- "mot_phan" : đúng nhưng thiếu ý hoặc chưa rõ
- "sai"      : sai hoặc không liên quan

Nhan_xet: chỉ ra điểm đúng/thiếu, nhắc lại kiến thức trọng tâm (2-3 câu)."""

    raw = hoi_gia_su(prompt)
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

    try:
        data = json.loads(raw)
        ket_qua  = data.get("ket_qua", "sai")
        nhan_xet = data.get("nhan_xet", raw)
    except (json.JSONDecodeError, ValueError):
        ket_qua  = "dung" if any(kw in raw.lower() for kw in
                    ["đúng rồi", "chính xác", "đúng!", "tốt lắm"]) else "sai"
        nhan_xet = raw if raw else "Không nhận được phản hồi từ AI."

    la_dung = ket_qua == "dung"

    db.add(aiLog(
        id_ngDung=user.id_ngDung,
        loaiHanhDong="do_bai",
        noiDungInput=req.cau_tra_loi,
        noiDungOutput=nhan_xet,
    ))
    db.commit()

    return {
        "nhan_xet":   nhan_xet,
        "la_dung":    la_dung,
        "ket_qua":    ket_qua,
        "dap_an_mau": dap_an_text,
        "goi_y":      goi_y_text,
    }


# ── Lưu kết quả ─────────────────────────────────────────────────
@router.post("/luu-ket-qua")
def luu_ket_qua(
    req: LuuKetQuaRequest,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tong    = req.tong_cau
    so_dung = req.so_cau_dung
    diem    = round((so_dung / tong) * 10, 1) if tong > 0 else 0
    xep     = 1 if diem >= 8 else 2 if diem >= 6.5 else 3 if diem >= 5 else 4

    ls = lichSuLamKT(
        id_baiKiemTra=None, id_ngDung=user.id_ngDung,
        diem=diem, xepLoai=xep,
        tg_batDau=datetime.utcnow(), tg_ketThuc=datetime.utcnow(),
    )
    db.add(ls)
    db.commit()

    return {
        "diem": diem, "so_dung": so_dung, "tong": tong,
        "xep_loai": ["", "Giỏi", "Khá", "Trung bình", "Yếu"][xep],
    }


# ── Các endpoint upload giữ nguyên ──────────────────────────────
@router.get("/tai-file-mau")
def tai_file_mau():
    path = "/tmp/cau_hoi_do_bai_mau.xlsx"
    _tao_excel_mau(path)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="cau_hoi_do_bai_mau.xlsx",
    )


def _tao_excel_mau(path: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise RuntimeError("pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Câu hỏi dò bài"
    headers_row = ["chu_de", "tieu_de", "cau_hoi", "dap_an_mau", "goi_y", "do_kho", "thu_tu"]
    for col, h in enumerate(headers_row, 1):
        ws.cell(row=1, column=col, value=h)
    wb.save(path)


@router.post("/upload/excel")
async def upload_excel(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ten = (file.filename or "").lower()
    if not any(ten.endswith(e) for e in [".xlsx", ".xls", ".csv"]):
        raise HTTPException(400, "Chỉ chấp nhận .xlsx, .xls, .csv")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File quá lớn (tối đa 5MB)")
    try:
        kq = xu_ly_excel(content, file.filename, db)
    except ValueError as e:
        raise HTTPException(422, str(e))
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    return {"message": f"Đã thêm {kq['da_them']} câu hỏi", **kq}


@router.post("/upload/pdf")
async def upload_pdf(
    file: UploadFile = File(...),
    ten_chu_de: str = Form(...),
    tieu_de: str = Form(...),
    so_cau: int = Form(5),
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(400, "Chỉ chấp nhận .pdf")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File quá lớn (tối đa 10MB)")
    try:
        kq = xu_ly_pdf(content, ten_chu_de, tieu_de, max(3, min(10, so_cau)), db)
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"message": f"AI sinh {kq['da_them']} câu hỏi từ PDF", **kq}


@router.post("/upload/word")
async def upload_word(
    file: UploadFile = File(...),
    ten_chu_de: str = Form(...),
    tieu_de: str = Form(...),
    so_cau: int = Form(5),
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ten = (file.filename or "").lower()
    if not (ten.endswith(".docx") or ten.endswith(".doc")):
        raise HTTPException(400, "Chỉ chấp nhận .docx, .doc")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File quá lớn (tối đa 10MB)")
    try:
        kq = xu_ly_word(content, ten_chu_de, tieu_de, max(3, min(10, so_cau)), db)
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"message": f"AI sinh {kq['da_them']} câu hỏi từ Word", **kq}


@router.get("/chu-de")
def lay_chu_de(db: Session = Depends(get_db)):
    return [{"id": c.id_chuDe, "ten": c.ten_chuDe}
            for c in db.query(chuDe).order_by(chuDe.id_chuDe).all()]


@router.get("/bai-hoc/{id_chu_de}")
def lay_bai_hoc(id_chu_de: int, db: Session = Depends(get_db)):
    bais = db.query(taiLieu).filter(taiLieu.id_chuDe == id_chu_de).order_by(taiLieu.mucDoKho).all()
    return [{"id": b.id_taiLieu, "tieu_de": b.tieuDe, "loai": b.loai, "do_kho": b.mucDoKho} for b in bais]


@router.get("/cau-hoi/{id_tai_lieu}")
def lay_cau_hoi(id_tai_lieu: int, db: Session = Depends(get_db)):
    caus = db.query(cauHoi).filter(
        cauHoi.id_taiLieu == id_tai_lieu,
        cauHoi.id_loaiCauHoi == ID_DO_BAI
    ).order_by(cauHoi.thuTu).all()
    if not caus:
        raise HTTPException(404, "Chưa có câu hỏi cho bài này")
    return [{"id": c.id_cauHoi, "cau_hoi": c.noiDung, "thu_tu": c.thuTu} for c in caus]